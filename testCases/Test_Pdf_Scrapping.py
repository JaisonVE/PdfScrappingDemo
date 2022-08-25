from PyPDF2 import PdfFileReader
import openpyxl


file = "C:\\Users\\user\\PycharmProjects\\PdfScrapping\\TestData\\Readers Digest.xlsx"

workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"]

file_path = "C:\\Users\\user\\PycharmProjects\\PdfScrapping\\TestData\\Reader Digest.pdf"
pdf = PdfFileReader(file_path)
no_of_pages = len(pdf.pages)

for x in range(0, no_of_pages):
    page = pdf.pages[x]
    page_content = page.extract_text()
    r = x + 2
    p = x + 1
    sheet.cell(1,1).value = "Lines"
    sheet.cell(1,2).value = "Content"
    sheet.cell(r, 1).value = p
    sheet.cell(r, 2).value = page_content
    workbook.save(file)
